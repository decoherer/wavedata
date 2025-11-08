import pandas as pd
import numpy as np
from numpy import sqrt,pi,nan,inf,sign,exp,log,sin,cos,floor,ceil

def supresserror(f,default=None):
    try:
        return f()
    except Exception:
        return default
class Vec(pd.Series):
    """
    Extends pandas Series to apply methods to each element using dot notation.
    If the method does not exist on the element, a numpy or global function with the same name is applied.
    >>> Vec([-1, 2, -3]).abs() -> Vec([1, 2, 3])
    >>> Vec(['a', 'b', 'C']).upper() -> Vec(['A', 'B', 'C'])
    >>> Vec([1, 2, 3]).pow(Vec([2, 3, 2])) -> Vec([1, 8, 9])
    """
    # def __getattr__(self, name):
    #     def method(*args):
    #         aa = [a.values if isinstance(a, Vec) else a for a in args]
    #         return Vec([getattr(item, name)(*aa) if hasattr(item, name) else 
    #                     getattr(np, name)(item, *aa) if hasattr(np, name) else 
    #                     globals()[name](item, *aa) for item in self])
    #     return method
    def __getattr__(self, name):
        return lambda *args, **kwargs: Vec([getattr(item, name)(*args, **kwargs) if hasattr(item, name) else getattr(np, name)(item, *args, **kwargs) if hasattr(np, name) else globals()[name](item, *args, **kwargs) for item in self])
def maplistargs(func):
    # given f(a,b,c), check if a, b, or c is a list and if so return a list of results, one for each element
    # e.g. maplistargs(f)(a=[0,1,2],b=8,c=[10,11,12]) returns [f(0,8,10),f(1,8,11),f(2,8,12)]
    def wrapper(*args, **kwargs):
        argnames = func.__code__.co_varnames[:func.__code__.co_argcount] # combine args and kwargs into a single dict for easier processing
        calldict = dict(zip(argnames, args))
        calldict.update(kwargs)
        listargs = {k:v for k,v in calldict.items() if isinstance(v, list)} # only the ones that are lists
        if not listargs:  # No list arguments, call the function directly
            return func(*args, **kwargs)
        vals = list(listargs.values())
        if not all(len(v) == len(vals[0]) for v in vals):
            raise ValueError("All list arguments must have the same length.")
        keys    = tuple(listargs.keys())
        vals    = tuple(listargs[k] for k in keys)
        nonlist = {k:v for k,v in calldict.items() if k not in listargs}
        allargs = [{**nonlist, **dict(zip(keys, tup))} for tup in zip(*vals)]
        return [func(**d) for d in allargs]
    return wrapper

# def maplistargs(func):
#     # given f(a,b,c), check if a, b, or c is a list and if so return a list of results, one for each element
#     # e.g. maplistargs(f)(a=[0,1,2],b=8,c=[10,11,12]) returns [f(0,8,10),f(1,8,11),f(2,8,12)]
#     def wrapper(*args):
#         print('args',args)
#         list_args = [isinstance(arg, list) for arg in args]
#         if any(list_args):
#             results = []
#             list_length = None
#             for arg in args:
#                 if isinstance(arg, list):
#                     if list_length is None:
#                         list_length = len(arg)
#                     elif len(arg) != list_length:
#                         raise ValueError("All list arguments must have the same length.")
#             for i in range(list_length):
#                 current_args = [arg[i] if isinstance(arg, list) else arg for arg in args]
#                 results.append(func(*current_args))
#             return results
#         else:
#             return func(*args)
#     return wrapper
def storecallargs(func):
    # Captures all positional and keyword arguments passed to the decorated function.
    # Stores these captured arguments and keyword arguments in self.callargs as a dictionary with two keys: 
    # 'args' for a tuple of positional arguments and 'kwargs' for a dictionary of keyword arguments.
    # see storecallargstest() for an example
    from functools import wraps
    @wraps(func)
    def wrapper(self, *args, **kwargs):
        arg_names = func.__code__.co_varnames[1:func.__code__.co_argcount]  # Skip 'self'
        callargs = dict(zip(arg_names, args))
        callargs.update(kwargs)
        result = func(self, *args, **kwargs)
        self.callargs = callargs
        return result
    return wrapper
def dumpargs(func): # Decorator to print function call details # https://stackoverflow.com/a/6278457
    import inspect
    from functools import wraps
    @wraps(func)
    def wrapper(*args, **kwargs):
        func_args = inspect.signature(func).bind(*args, **kwargs).arguments
        func_args_str = ", ".join(map("{0[0]}={0[1]!r}".format, func_args.items()))
        f = func(*args, **kwargs)
        # print(f"{func.__module__}.{func.__qualname__}({func_args_str})")
        print(f"{func.__name__}({func_args_str}) = {f}") 
        return f
    return wrapper
def montecarlo(f,*xx,n,xnames=None,verbose=True,plot=False): # n = # of iterations, xnames = name of each dim
    # find max of an N-dim function where N is the number of arrays in xx and each array is a 1-dim array of function arguments
    # e.g. xx = xs,ys,zs = (-1,0,+1),(0,1,2),(0,3,6,9)
    N,Ms = len(xx),[len(list(a)) for a in xx] # Ms = length of each arg list
    def val():
        from random import choice
        return [choice(xi) for xi in xx]
    def ff(i,*x):
        if verbose: print(f"{i:8d},  ({list2str(x,sep=',')}),",end='')
        try:
            result = f(*x)
            if verbose: print(f'{result:g}')
        except Exception as e:
            result = None
            if verbose: print(e)
        return result
    vs = [val() for _ in range(n)]
    fs = [ff(i,*v) for i,v in enumerate(vs)]
    fs,vs = zip(*sorted([(fi,v) for fi,v in zip(fs,vs) if fi is not None]))
    if plot:
        for i in range(N):
            xys = sorted([(v[i],fi) for v,fi in zip(vs,fs)])
            from wavedata import Wave
            name = xnames[i] if xnames else ''
            Wave.fromxys(xys).plot(m='o',grid=1,c=str(i),seed=0,x=name)
    yxs = [(fi,v) for fi,v in zip(fs,vs)]
    if verbose:
        print('top twenty:')
        for y,x in yxs[-20:]: print(y,x)
    ftop,vtop = yxs[-1]
    return vtop,ftop

def stepclimb(f,*xx,start=None,startval=None,cornersearch=False,res=0,verbose=True): # res = upsample each x in xx by 2**res
    from math import isnan
    def upsample(a,n):
        from wavedata import Wave
        return upsample(Wave(a).upsample(num=2).ylist(),n-1) if 0<n else a
    xx = [upsample(a,res) for a in xx]
    # find max of an N-dim function where N is the number of arrays in xx and each array is a 1-dim array of function arguments
    # e.g. xx = xs,ys,zs = (-1,0,+1),(0,1,2),(0,3,6,9)
    # search will start at the center of the grid by default, e.g. f(x=0,y=1,z=6)
    # or choose start='random' for random grid index
    # specify start=[i,j,k] by index or startval=(x,y,z) by value
    N,Ms = len(xx),[len(list(a)) for a in xx]
    def g(*ii): # g(i,j,k) == f(xs[i],ys[j],zs[k])
        g.cache = {} if not hasattr(g,'cache') else g.cache
        if ii in g.cache:
            return g.cache[ii]
        v = f(*[a[i] for i,a in zip(ii,xx)])
        v = -np.inf if (v is None or isnan(v)) else v
        g.cache[ii] = v
        return v
    def hillclimb1d(ii,dim):
        i0,M = ii[dim],Ms[dim]
        def k2ii(k):
            return [(k if d==dim else i) for d,i in enumerate(ii)]
        def gg(k):
            return g(*k2ii(k))
        def scan1d(dir=+1,n=0): # returns index of highest found
            ks = range(i0,M if +1==dir else -1,dir)
            n,k = 0,i0
            for n0,(k0,k1) in enumerate(zip(ks[:-1],ks[1:])):
                if gg(k0)>gg(k1):
                    n,k = n0+1,k0
                    break
                n,k = n0+1,k1
            # if verbose: print(f'      | {gg(k):g} dim {dim} dir {dir:+d} k {k} ks',list(ks[:n]))
            if verbose: print(f'      | {gg(k):g} dim {dim} dir {dir:+d} ({list2str(ii2vals(k2ii(k)))})')
            return k
        k = scan1d(+1)                      # scan for higher value in +1 direction
        k = scan1d(-1) if k==i0 else k      # if higher not found scan in -1 direction
        return k2ii(k)
    def centerii():
        from random import choice
        # return [choice(M) for M in Ms]
        return [choice(range(M)) for M in Ms]
    def ii2vals(ii):
        return [a[i] for i,a in zip(ii,xx)]
    def nearestindex(a,x):
        dxs = [abs(ai-x) for ai in a]
        return sorted([(dx,i) for i,dx in enumerate(dxs)])[0][1]
    def getindex(a,x): # equivalent version of list(a).index(x) using np.close
        jj = np.flatnonzero(np.isclose(a,x))
        return None if 0==len(jj) else int(jj[0])
    def vals2ii(vals,findnearest=True):
        return [(nearestindex if findnearest else getindex)(a,v) for v,a in zip(vals,xx)]
    ii = ii0 = vals2ii(startval) if startval is not None else centerii() if start is None else randomii() if 'r'==start[0] else start
    if verbose and (start is None or 'r'==start[0]):
        print('start',ii,tuple([a[i] for i,a in zip(ii,xx)]),'|',g(*ii))
    while 1: # step along direct neighbors
        for dim in range(N):
            ii1 = hillclimb1d(ii,dim)
            assert g(*ii1)>=g(*ii), f'new{g(*ii1):g}<old{g(*ii):g}'
            ii = ii1
        if ii==ii0:
            break
        ii0 = ii
    def neighborclimb(ii0):
        # check all neighbors
        # ( 6 direct neighbors and  20 corner neighbors for dim=3)
        # ( 8 direct neighbors and  72 corner neighbors for dim=4)
        # (10 direct neighbors and 232 corner neighbors for dim=5)
        # (2n direct neighbors and 3ⁿ-2n-1 corner neighbors for dim=n)
        def neighbor(ii,dii):
            return tuple(i+di for i,di in zip(ii,dii))
        def dist(ii):
            return sum([abs(i) for i in ii])
        def valid(ii):
            return all(0<=i<M for i,M in zip(ii,Ms))
        from itertools import product
        diis = sorted(product((-1,0,+1),repeat=N),key=dist)[1:] # sorted by distance to corner
        neighbors = list(filter(valid,[neighbor(ii0,dii) for dii in diis]))
        # print('diis',diis); print('neighbors',[neighbor(ii0,dii) for dii in diis]); print('valid neighbors',neighbors)
        g0 = g(*ii0)
        if verbose: print(f"      ({list2str(ii2vals(ii0))}), {g0:g}")
        for ii in neighbors:
            if verbose: print(f"      ({list2str(ii2vals(ii0))}) → ({list2str(ii2vals(ii))})")
            if g(*ii0)<g(*ii):
                return ii
        return ii0
    while cornersearch:
        ii = neighborclimb(ii0)
        if ii0==ii:
            break
        ii0 = ii
    xx0 = tuple([a[i] for i,a in zip(ii,xx)])
    if verbose: print(f"      | final: ({list2str(xx0,sep=',')}), {g(*ii):g}")
    return ii,xx0,g(*ii)
def fileformat(file,delimiter=None,crop=False,gethead=False):
    def convert(line):
        def string2symbol(s):
            v = s.strip()
            if not v:
                return ' '
            try:
                float(v)
                return '0'
            except ValueError:
                return 'x'
        return ','.join([string2symbol(s) for s in line.split(delimiter)])
    with open(file, 'r') as f:
        text = f.read()
    delimiter = delimiter if delimiter is not None else ',' if text.count(',')>text.count('\t') else '\t'
    lines = text.rstrip('\n').split('\n')
    out = [convert(line) for line in lines]
    n = len(out)
    while 1<n and out[n-1]==out[n-2]:
        n -= 1
    assert '0' not in out[n-2], 'header incorrectly found' # print(out[n-2])
    if gethead:
        return n-2,delimiter,out
    return '\n'.join(out[:n] if crop else out)
def loadfile(file,delimiter=None,quiet=False,debug=False,waves=False):
    n,d,ff = fileformat(file,delimiter=delimiter,gethead=True)
    with open(file, 'r') as f:
        rows = [row.strip('\n') for row in f.readlines()]
    def sanitize(s):
        s = s.replace('%','pct')
        s = s.replace('*','x').replace('/','per')
        s = s.replace('(','_').replace(')','')
        s = s.replace('[','_').replace(']','')
        return s.strip().replace(' ','')
    names = [sanitize(s) for s in rows[n].strip(' #'+d).split(d)]
    if debug:
        print(f'\nhead row:',n,f'\ndelimiter: "{d}"',f'\nhead:',ff[n],f'\ndata:',ff[n+1],f'\nnames',names,f'\nhead:',rows[n],f'\ndata:',rows[n+1])
    values = [[float(s) for s in row.strip(' '+d).split(d)] for row in rows[n+1:]]
    if not quiet:
        if not len(names)==len(values[0]):
            print(f'mismatch headcount:{len(names)} datacount:{len(values[0])}')
        print(f'{len(names)} columns loaded:',list2str(names,f='{}',sep=','))
    aa = [np.array(a) for a in transpose(values)]
    if waves:
        from wavedata import Wave
        return [Wave(y,aa[0]) for y in aa[1:]] if waves else (names,aa)
    return (names,aa)
def csvcolumns(file,cols=None,skip=0,endskip=0,delimiter=','):
    import csv
    with open(file, 'r') as f:
        rows = [row for row in csv.reader(f,delimiter=delimiter)][skip:]
    rows = rows[:len(rows)-endskip]
    cols = cols if cols is not None else range(max([len(row) for row in rows]))
    def column(i):
        return [float(row[i].strip().strip('\ufeff')) for row in rows]
    return [np.array(column(i)) for i in cols]
def excelcolumn(file,column='A',rowstart=1,rowend=None,sheet=0,filtertext=1): # sheets are zero-indexed, cells are one-indexed
    import openpyxl
    def convertcsv(file):
        import csv
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(file) as f:
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)
        newfile = file.replace('.csv','.xlsx')
        wb.save(newfile)
        return newfile
    def str2num(x):
        try:
            return float(x)
        except (ValueError,TypeError):
            return None
    file = convertcsv(file) if file.endswith('.csv') else file
    # print('file',file)
    def lstripstr(a): # filter out strings and None values from the start of a list # gpt4
        return next((a[i:] for i,ai in enumerate(a) if not (isinstance(ai,str) or ai is None)), [])
    def stripstr(a):
        return lstripstr(lstripstr(a[::-1])[::-1])
    import openpyxl
    # import xlrd # only for old .xls
    wb = openpyxl.load_workbook(file,data_only=True) # use data_only=False to get cell formula e.g. '=A2+1'
    ws = wb.worksheets[sheet] # print('ws',ws)
    # first_sheet = wb.get_sheet_names()[0]
    # worksheet = wb.get_sheet_by_name(first_sheet)
    rowend = (ws.max_row+rowend if rowend<0 else rowend) if rowend is not None else ws.max_row
    col = [ws[column+f'{n}'].value for n in range(rowstart,rowend+1)] # print('col',col) # print('ws[D19]',ws['D19'].value) # 
    if filtertext:
        strip = stripstr if 1<filtertext else lstripstr
        out = strip(col) if strip(col) else [str2num(s) for s in col]
        return np.array([x if x is not None else nan for x in out]) # print('col',col)
    return col
def loadvna(file,folder='',sheet=0):
    def load(col):
        if file.endswith('csv'):
            ys,xs = csvcolumns(folder+file,cols=(col,0),skip=8,endskip=2)
        else:
            xs = excelcolumn(folder+file,column='A',rowstart=8,rowend=-1,sheet=sheet,filtertext=1)
            ys = excelcolumn(folder+file,column='ABCDEFGHI'[col],rowstart=8,rowend=-1,sheet=sheet,filtertext=1)
        from wavedata import Wave
        return Wave(ys,1e-9*xs)
    from waves import csvcolumns
    s11 = load(col=1).rename('S11')
    s12 = load(col=3).rename('S12')
    s21 = load(col=5).rename('S21')
    s22 = load(col=7).rename('S22')
    return s11,s12,s21,s22
def nploadfile(file,dtype=None,delimiter='\t',skip=0):
    return np.array(np.genfromtxt(file, dtype=dtype, delimiter=delimiter, names=True, skip_header=skip).tolist())
def npsavecsv(arrays,names,names2=None,filename='tmp.csv',delimiter=','):
    with open(filename, 'w') as f:
        f.write(delimiter.join(names)+'\n')
        if names2 is not None:
            f.write(delimiter.join(names2)+'\n')
        for row in zip(*arrays):
            f.write(delimiter.join([str(x) for x in row])+'\n')
        
def track(iter,f="{:g}",end=' '): # usage: [f(n) for n in track(ns)] instead of [f(n) for n in ns]
    for i in iter:
        try:
            print(i if hasattr(i,'__len__') else f.format(i),end=end)
        except:
            print('.',end=end)
        yield i
    print()
def batchtrack(a,batch=10):
    da = 10
    for k,i in enumerate(a):
        if k%da==0: print('.'+' '*((k//da+1)%10==0)+'\n'*((k//da+1)%100==0),end='')
        yield i
    print()
def percenttrack(a):
    # g = (s for s in '0123456789')
    da = len(a)/100
    for k,i in enumerate(a):
        # if k%(len(a)//10)==0: print(next(g),end='')
        if k%da==0: print('.'+' '*((k//da+1)%10==0)*((k//da+1)//10<10),end='')
        yield i
    print()
def list2str(a,f='{:g}',sep=' '): # e.g. print(list2str([1,2,3],f='{:.1f}',sep='#')) # 1.0#2.0#3.0
    if 0==len(a): return ''
    if hasattr(a[0],'__len__') and not isinstance(a[0],str):
        aa = [f"({list2str(ai,f,sep=',')})" for ai in a]
        return sep.join((len(aa)*['{}'])).format(*aa)
    return sep.join((len(a)*[f])).format(*a)
def dict2str(d,f='{}',sep=' '):
    ss = [f"{v}{k}" for k,v in d.items()]
    return sep.join(ss)
    # return sep.join(*(len(d)*'{}')).format(*ss)
def maximizeexample(disp=True):
    def f(x):
        x0,y0 = x
        g = exp(-(x0-np.sqrt(2))**2-(y0-np.sqrt(2))**2)
        return 1/g
    x0,y0 = (0,0)
    result = scipy.optimize.minimize(f, (x0,y0), bounds=((-2,+2),(-2,+2)), options={'disp':disp})
    x0,y0 = result.x
    print('x0,y0',x0,y0)
def loadcsv(file,verbose=False,**kwargs):
    import pandas as pd
    df = pd.read_csv(file,**kwargs)
    if verbose:
        print('df',df)
        print('df.shape',df.shape)
        print('df.index',df.index)
        print('df[df.columns[0]]',list(df[df.columns[0]]))
        print('df.columns',df.columns)
    return df
def finvert(f,x0,x1,maxiter=None,xtol=None,verbose=False):
    import scipy
    from functools import lru_cache
    @lru_cache
    def func(n):
        y = f(n)
        if verbose: print('   ',n,'→',y)
        return y
    solution = scipy.optimize.root_scalar(func, bracket=[x0,x1], method='brentq', maxiter=maxiter, xtol=xtol)
    if verbose:
        print('   iterations:',solution.iterations,'   calls:',solution.function_calls,'   solution:',solution)
    return solution.root
def findxgivenf(f,x0,x1,dx=None,maxiter=None,dxonlast=True): # find x0<x<x1 such that f(x)=0, must have f(x0)*f(x1)<0
    import scipy
    from functools import lru_cache
    @lru_cache
    def func(n):
        return f(n)
    if dx is None:
        solution = scipy.optimize.root_scalar(func, bracket=[x0,x1], method='brentq', maxiter=maxiter)
        # print(solution)
        return solution.root
    y0,y1 = func(float(x0)),func(float(x1))
    def roundx(x):
        return round(int(x/dx))*dx
    x = roundx(lerp(0,y0,y1,x0,x1))
    y = func(float(x))
    xx = roundx(x0/2+x/2) if y*y0<0 else roundx(x/2+x1/2)
    if xx==x or (maxiter is not None and maxiter<=1):
        if dxonlast:
            return x
        x,xx = (x0,xx) if y*y0<0 else (xx,x1)
        assert abs(round((xx-x)/dx))==1. # print('x,xx',x,xx)
        return findxgivenf(func,x0=x,x1=xx,dx=None,maxiter=1)
    yy = func(float(xx))
    # print('y0,yy,y,y1',y0,yy,y,y1); print('x0,xx,x,x1',x0,xx,x,x1); print()
    if y*yy>0:
        x,xx = (x0,xx) if y*y0<0 else (xx,x1)# print('x,xx',x,xx)
    # assert y*yy<=0, 'not yet implemented if not y<0<yy or yy<0<y'
    if not func(float(x))*func(float(xx))<=0: raise NotImplementedError
    return findxgivenf(func,xx,x,dx=dx,maxiter=(None if maxiter is None else maxiter-1),dxonlast=dxonlast)
def piecewise(x,xs,fs): # easy to use version of np.piecewise
    # xs = region boundaries
    # fs = function to use in each region
    assert len(xs)+1==len(fs)
    if 0==len(xs):
        return fs[0](x)
    return np.where(x < xs[0], fs[0](x), piecewise(x,xs[1:],fs[1:]))
# Wave(piecewise(np.linspace(0,4,1001),[1,2,3],[lambda x:-x**2,np.sin,lambda x:x,np.cos])).plot()
def pickmin(vals): # return index of min
    return min(range(len(vals)), key=vals.__getitem__)
def pickmax(vals): # return index of max
    return max(range(len(vals)), key=vals.__getitem__)
def getindex(a,x): # equivalent version of list(a).index(x) using np.close
    jj = np.flatnonzero(np.isclose(a,x))
    return None if 0==len(jj) else int(jj[0])
def transpose(listoflists): # returns list of tuples
    return [ll for ll in zip(*listoflists)]
def deltas(A):
    return [b-a for b,a in zip(A[1:],A[:-1])]
def headtail(A,n=3):
    return A[:n],A[-n:]
def nths(A,n):
    return [x[n] for x in A]
def nthitem(gen,n):
    from itertools import islice
    return next(islice(gen, start=n, stop=None))
def listmiddle(A):
    N = len(A)
    assert 1==N%2
    return A[N//2]
def listmiddles(A,n=None):
    N = len(A)
    n = n if n is not None else (1 if N%2 else 2)
    assert n%2==N%2
    return A[(N+1-n)//2:(N+1+n)//2]
def listends(A):
    return list(A[:1])+list(A[-1:])
def cutmid(A):
    return (A[0],A[-1])
def middle(n,N): # return the n values that are in the middle of range(N)
    assert n%2==N%2
    return list(range((N+1-n)//2,(N+1+n)//2))
def extend(A,n=1):
    (a,b),(y,z) = A[:2],A[-2:]
    B = [2*a-b,*A,2*z-y]
    return B if n<=1 else extend(B,n-1)
def fixfloatdrift(x,tol=1e-6):
    return type(x)([fixfloatdrift(xi,tol) for xi in x]) if hasattr(x,'__len__') else round(x/tol)*tol
def hrange(x0,x1,n):
    # split range into n equal segments, returning the x position of centers of each segment (useful for histograms)
    dx = (x0-x1)/n
    return np.linspace(x0+dx/2,x1-dx/2,n)
def frange(x0,x1,dx,z,dz,Δz,tol=1e-6,reversed=False):
    # focused range, lowers the resolution from dx to dz within Δz range around z (where x0<z<x1)
    assert x0<=z<=x1, f"frange requires x0<z<x1 x0:{x0:g} z:{z:g} x1:{x1:g}"
    raise NotImplementedError
def error(s):
    assert 0, s
def logrange(x0,x1,res=1):
    if x1<x0:
        return logrange(x1,x0,res=res)[::-1]
    def highreslevels(dx,dy,dz):
        return wrange(1,2,dx,array=0,endpoint=0) + wrange(2,5,dy,array=0,endpoint=0) + wrange(5,10,dz,array=0,endpoint=1)
    levels = [1,2,5,10] # for i in range(res-1): levels = sorted(levels+[0.5*a+0.5*b for a,b in zip(levels[:-1],levels[1:])])
    levels = levels if res<2 else [1,1.4,2,3,5,7,10]
    levels = levels if res<3 else [1.,1.2,1.4,1.6,1.8,2.,2.5,3.,3.5,4.,4.5,5.,6.,7.,8.,9.,10.]
    levels = levels if res<3 else highreslevels(0.2,0.5,1)
    levels = levels if res<4 else highreslevels(0.1,0.2,0.5)
    levels = levels if res<5 else highreslevels(0.05,0.1,0.2)
    levels = levels if res<6 else highreslevels(0.02,0.05,0.1)
    levels = levels if res<7 else error(f"res={res} not yet implemented")
    e0,e1 = np.floor(np.log10(x0)),np.ceil(np.log10(x1))+1
    a = [x*10**n for n in wrange(e0,e1) for x in levels[:-1]]
    return np.array([ai for ai in a if x0<=ai<=x1])
def logrounddown(x):
    f = np.floor(np.log10(x))
    c = max([z for z in [1,2,5] if z<=x/10**f])
    # print(x,10**f,x/10**f,c,c * 10**f)
    return c * 10**f
def wrange(x0,x1,dx=1,endround=False,endpoint=True,wave=False,array=True,aslist=False,tol=1e-9,reverse=False,format='{:g}'):
    x0,x1 = (round(x0/dx)*dx,round(x1/dx)*dx) if endround else (x0,x1)
    # assert x0<=x1, f"wrange requires x0<x1 x0:{x0:g} x1:{x1:g}"
    if x1<x0:
        return wrange(x1,x0,dx=dx,endround=endround,endpoint=endpoint,wave=wave,array=array,aslist=aslist,tol=tol,reverse=reverse)[::-1]
    n = int(round((x1-x0)/dx))
    assert abs(x1-x0-n*dx) < tol*dx, f'wrange invalid spacing given tolerance, x1-x0:{x1-x0} dx:{dx} n:{n} x1-x0-n*dx:{abs(x1-x0-n*dx)}'
    # return evenly spaced points between x0 and x1 inclusive, with closest step possible to dx
    xs = np.array([float(format.format(x)) for x in np.linspace(x0,x1,n+1)]) if format else np.linspace(x0,x1,n+1)
    assert np.abs(xs-np.linspace(x0,x1,n+1)).max()<tol, f'wrange failed to round to tolerance {tol}: {tol}<{xs-np.linspace(x0,x1,n+1)}'
    xs = xs if endpoint else xs[:-1]
    from wavedata import Wave
    return Wave(xs,xs) if True==wave else xs if (array and not aslist) else list(xs)
def widths2grid(ds,x0=None):
    assert all(0<d for d in ds)
    x0 = x0 if x0 is not None else -sum(ds)/2
    return x0 + np.nancumsum([0]+list(ds))
def rint(x):
    return int(np.round(x))
def funcstring():
    import inspect
    frame = inspect.currentframe().f_back
    name = inspect.getframeinfo(frame).function
    arginfo = inspect.getargvalues(frame)
    vals = [arginfo.locals[a] for a in arginfo.args]
    # print( inspect.getargvalues(frame) )
    return name+''.join([f' {v}{k}' for k,v in zip(arginfo.locals,vals)])
def dargs():
    import inspect
    frame = inspect.currentframe().f_back
    arginfo = inspect.getargvalues(frame)
    vals = [arginfo.locals[a] for a in arginfo.args]
    return {k:v for k,v in zip(arginfo.locals,vals)}
def funcname(info=False):
    import inspect
    frameinfo = inspect.getframeinfo(inspect.currentframe().f_back)
    if info: return frameinfo
    return frameinfo.function
def savestring(**d):
    return ' '.join([f"{d[v]:g}{v}" for v in d])
def default(x,x0):
    return x0 if x is None else x
def lacc(*args,**kwargs):
    from itertools import accumulate
    return list(accumulate(*args,**kwargs))
def lmap(*args):
    return list(map(*args))
def lrange(*args):
    return list(range(*args))
def lfilter(*args):
    return list(filter(*args))
def deal(n,aa,reverse=False):
    if reverse:
        return [np.array(aa[::-1][i::n]) for i in range(n)][::-1]
    return [np.array(aa[i::n]) for i in range(n)]
deal2,deal3,deal4,deal5 = [(lambda a,r=False,n=i : deal(n,a,r)) for i in [2,3,4,5]]
def autodeal(aa,aswaves=False):
    def ismonotonicincreasing(aa):
        return all(a<b for a,b in zip(aa,aa[1:]))
    for n in range(2,6):
        if len(aa)%n==0 and ismonotonicincreasing(aa[::n]):
            if aswaves:
                x,*ys = deal(n,aa)
                from wavedata import Wave
                return [Wave(y,x) for y in ys]
            return deal(n,aa)
    assert 0, f"autodeal failed to find 2<=columns<=6"
def removecloseduplicates(a,**args):
    if len(a)<2:
        return a
    if np.isclose(a[0],a[1],**args):
        return removecloseduplicates(a[1:],**args)
    return np.append(a[:1], removecloseduplicates(a[1:],**args))
def gauss(x,*p):
    a,y0,fwhm,x0 = p
    return y0+a*np.exp(np.log(0.5)*(x-x0)**2/(fwhm/2.)**2)
def gausscurvefit(xs,ys,plot=False,x0=None,sdev=None):
    from scipy.optimize import curve_fit
    if x0 is None:
        p0 = (np.max(ys),np.mean(ys),5,xs[ys.argmax()]) # initial guess for Gaussian
        p,pcov = curve_fit(gauss, xs, ys, p0=p0, sigma=sdev)
        psdev = np.sqrt(np.diag(pcov))
        height,floor,fwhm,x0 = p
        yys = gauss(xs,*p)
    else:
        def fixedgauss(x,*p):
            a,y0,fwhm = p
            return y0+a*np.exp(np.log(0.5)*(x-x0)**2/(fwhm/2.)**2)
        p0 = (np.max(ys),np.mean(ys),5)
        p,pcov = curve_fit(fixedgauss, xs, ys, p0=p0, sigma=sdev)
        psdev = np.sqrt(np.diag(pcov))
        height,floor,fwhm = p
        yys = fixedgauss(xs,*p)
    if plot: # convert back to total counts (not per sec) just for the plot
        plt.plot(xs,ys,'darkred',xs,yys,'darkblue')
        plt.show() #plt.savefig('h12.png') #if plot: os.startfile('h12.png')
    #print('guess:',p0,'fit:',p,'sdev:',psdev)
    return height,floor,fwhm,x0
def discretedeltafunction(x,x0,dx): # constant area=1
    return np.maximum(0,1-np.abs((x-x0)/dx))/dx
def nptophat(x,x0,x1):
    return np.heaviside(x-x0,0.5)-np.heaviside(x-x1,0.5)
def stepfunction(x,x0,dx):
    # return np.clip((x-x0)/dx,0,1)
    return np.clip(0.5*(x-x0+dx)/dx,0,1)
def inversestepfunction(x,x0,dx):
    return stepfunction(-x,-x0,dx)
def tophat(x,x0,x1,dx):
    # guaranteed constant area=x1-x0 on dx spaced grid
    # also guaranteed that tophat(xs,x0,x1,dx) + tophat(xs,x1,x2,dx) = tophat(xs,x0,x2,dx)
    # tri = 0.5 + 0.5*(x1-x0)/dx - np.abs(x-0.5*(x0+x1))/dx
    # return np.maximum(0,np.minimum(1,tri))
    return stepfunction(x,x0,dx) - stepfunction(x,x1,dx)
def sinc(x):
    return np.sinc(x/np.pi)
def sincsqr(x,*p):
    a,y0,fwhm,x0 = p
    sincsqrhwhm = 1.39155737825151004649626429454656317830085754394531250
    assert np.allclose(0.5,sinc(sincsqrhwhm)**2)
    # def sinc(x): return np.sinc(x/np.pi)
    # assert np.allclose(np.sin(1)/1,sinc(1)) and np.allclose(np.sin(0.3)/0.3,sinc(0.3))
    return y0+a*np.sinc(2*sincsqrhwhm*(x-x0)/fwhm/np.pi)**2 # doesn't work when x is np.array?
def sincsqrcurvefit(xs,ys):
    from scipy.optimize import curve_fit
    p0 = (np.max(ys),np.min(ys),(xs[-1]-xs[0])/10,xs[ys.argmax()]) # initial guess for Gaussian (fwhm guess is 10% scan range)
    p,pcov = curve_fit(sincsqr, xs, ys, p0=p0)
    psdev = np.sqrt(np.diag(pcov))
    height,floor,fwhm,x0 = p
    yys = sincsqr(np.array(xs),*p)
    return height,floor,abs(fwhm),x0
def cosfitfunc(x,*p):
    a,y0,period,x0 = p
    return y0+a*np.cos(2*np.pi*(x-x0)/period)
def coscurvefit(xs,ys,guess=(None,None,None,None),fix='0000'):
    p0 = (np.max(ys)-np.min(ys),np.min(ys),(xs[-1]-xs[0])/2,xs[ys.argmax()]) # initial default guesses # default period guess is 50% scan range
    p0 = tuple([(g if g is not None else p) for p,g in zip(p0,guess)])
    p = curvefit(xs,ys,cosfitfunc,guess=p0,fix=fix)
    a,y0,period,x0 = p
    yys = cosfitfunc(np.array(xs),*p)
    return a,y0,abs(period),x0%abs(period)
def quadfit(xs,ys,guess=(None,None,None),fix='000'):
    if (None,None,None)==guess:
        # x0 = np.mean(xs)
        # a = 4*(np.max(ys)-np.min(ys))/(np.max(xs)-np.min(xs))
        # guess = (a,-2*a*x0,a*x0**2)
        return np.polyfit(xs,ys,2)
    def fitfunc(x,*p):
        a,b,c = p
        return a*x**2 + b*x + c
    p = curvefit(xs,ys,fitfunc,guess=guess,fix=fix)
    a,b,c = p
    yys = fitfunc(np.array(xs),*p)
    return a,b,c
def polyfit(xs,ys,n,guess=None,fix=None):
    if None==guess:
        return np.polyfit(xs,ys,n)
    fix = fix if fix is not None else (n+1)*'0'
    def fitfunc(x,*p):
        return np.polyval(p,x)
    p = curvefit(xs,ys,fitfunc,guess=guess,fix=fix)
    yys = fitfunc(np.array(xs),*p)
    return p
def curvefit(xs,ys,fitfunc,guess,fix=None,maxfev=0,debug=False):
    # guess = list of coefficients (c0,c1,...) for fitfunc(x,c0,c1,...)
    # fix = '0110' or (0,1,1,0), for example, to mark coefficients not to vary
    # maxfev = max number of function evaluations
    fix = fix if fix is not None else [False for g in guess]
    from scipy.optimize import curve_fit
    assert len(guess)==len(fix) and all([str(b) in ('0','1','True','False') for b in fix])
    if all([str(b) in ('1','True','False') for b in fix]):
        return guess
    n,fix = len(guess),list(map(int,fix)) if isinstance(fix,str) else fix
    ifit = [i for i in range(n) if not fix[i]]
    def reducedcoef(p):
        return tuple([p[i] for i in range(n) if i in ifit])
    def fullcoef(q):
        d = {ifit[j]:j for j in range(len(ifit))}
        return tuple([q[d[i]] if i in ifit else guess[i] for i in range(n)])
    def f(x,*q):
        return fitfunc(x,*fullcoef(q))
    q0 = reducedcoef(guess)
    xs,ys = zip(*[(x,y) for x,y in zip(xs,ys) if not np.isnan(x) and not np.isnan(y)])
    if debug:
        from wavedata import Wave
        print('guess ',guess); print('fullq0',fullcoef(q0),'q0',q0)
        Wave.plots( Wave(ys,xs).setplot(m='o',l=''), Wave([f(x,*q0) for x in xs],xs) )
    q,qcov = curve_fit(f, xs, ys, p0=q0, maxfev=maxfev)
    # psdev = np.sqrt(np.diag(qcov))
    return fullcoef(q)
def gaussianproduct(θ0,σ0,ρ0,θ1,σ1,ρ1,degrees=False):
    # product of exp(-x²/σ0²)exp(-y²/ρ0²) rotated by θ0 with exp(-x²/σ1²)exp(-y²/ρ1²) rotated by θ1, returns resulting θ,σ,ρ 
    def abc(θ,σ,ρ): # https://en.wikipedia.org/wiki/Gaussian_function#Meaning_of_parameters_for_the_general_equation
        a = 0.5*cos(θ)**2/σ**2 + 0.5*sin(θ)**2/ρ**2
        b = 0.25*sin(2*θ)/ρ**2 - 0.25*sin(2*θ)/σ**2
        c = 0.5*sin(θ)**2/σ**2 + 0.5*cos(θ)**2/ρ**2
        return (a,b,c)
    def θσρ(a,b,c):
        θ = 0.5*np.arctan2(2*b,a-c)
        assert -pi/2<=θ<=+pi/2, f"b{b:g} a-c{a-c:g} θ{θ:g}"
        # print(f"b:{b:g} a-c:{a-c:g} θ:{θ:g}")
        σ = np.sqrt(0.5/(a*cos(θ)**2 + 2*b*cos(θ)*sin(θ) + c*sin(θ)**2))
        ρ = np.sqrt(0.5/(a*sin(θ)**2 - 2*b*cos(θ)*sin(θ) + c*cos(θ)**2))
        # return (θ,σ,ρ)
        return normalize(θ,σ,ρ)
    def normalize(θ,σ,ρ):
        def θnorm(θ): # restrict to ±π, e.g. 135° → -45°
            return (θ+0.5*pi)%pi - 0.5*pi
        θ,σ,ρ = np.where(σ>ρ,(θnorm(θ),σ,ρ),(θnorm(θ+0.5*pi),ρ,σ))
        # θ,σ,ρ = np.where(np.allclose(σ,ρ)&(np.abs(θ)>=0.25*pi),(θnorm(θ+0.5*pi),ρ,σ),(θnorm(θ),σ,ρ))
        θ,σ,ρ = np.where(np.allclose(σ,ρ),(0,σ,σ),(θnorm(θ),σ,ρ))
        return θ,σ,ρ
    θ0,θ1 = (θ0*pi/180,θ1*pi/180) if degrees else (θ0,θ1)
    (a0,b0,c0),(a1,b1,c1) = abc(θ0,σ0,ρ0),abc(θ1,σ1,ρ1)
    θ,σ,ρ = θσρ(a0+a1,b0+b1,c0+c1)
    return (θ*180/pi if degrees else θ),σ,ρ
def trapezoidintegrate(ys,xs,invfunc=False,returnarray=False):
    # returns function f(x) that gives the exact trapezoidal integral area
    # by definition, f(x<xs[0])==0 and f(xs[-1]<x)==total area
    assert np.allclose(ys[0],ys[-1],0), f"{ys[0]:g},{ys[-1]:g} trapezoidintegrate untested for nonzero ends"
    assert all(0<=np.diff(xs))
    ΔAs = [0.5 * (ys[i+1]+ys[i]) * (xs[i+1]-xs[i]) for i in range(len(xs)-1)] # area of each trapezoid
    As = lacc(ΔAs,initial=0) # As[n] = cumulative area of first n trapezoids
    assert len(As)==len(xs)==len(ys)
    if returnarray:
        return As
    def f(x):
        i = sum([xi<x for xi in xs])-1
        if len(xs)-1==i:
            return As[-1]
        if i<0:
            return 0
        xi,Δx = xs[i],xs[i+1]-xs[i]
        yi,Δy = ys[i],ys[i+1]-ys[i]
        return As[i] + yi*(x-xi) + 0.5*Δy/Δx*(x-xi)**2
    # @np.vectorize # not working properly
    def finv(A):
        i = sum([Ai<=A for Ai in As])-1
        if len(As)-1==i:
            return xs[-1]
        if i<0:
            return xs[0]
        xi,Δx = xs[i],xs[i+1]-xs[i]
        yi,Δy = ys[i],ys[i+1]-ys[i]
        if 0==Δy:
            return (A-As[i])/yi+xi
        c = yi*Δx/Δy
        xp,xn = [g*sqrt(2*Δx/Δy*(A-As[i])+c**2)-c+xi for g in (+1,-1)]
        return xp if xi<=xp<=xi+Δx else xn
    return finv if invfunc else f
def discretepdfsample(xs,ys=None):
    # xs = choices, ys = probabilities
    # returns CDF⁻¹(x) with 0<x<1 where x=random.random() where CDF is the cumulative distribution function for the given PDF
    # https://en.wikipedia.org/wiki/Probability_density_function#Link_between_discrete_and_continuous_distributions
    ys = ys if ys is not None else [1 for x in xs]
    assert all([0<=y for y in ys])
    return np.random.choice(xs,p=[y/sum(ys) for y in ys])
def pdfsample(xs,ys=None,debug=False):
    # xs = choices, ys = probabilities
    # returns CDF⁻¹(x) with 0<x<1 where x=random.random() where CDF is the cumulative distribution function for the given PDF
    # https://en.wikipedia.org/wiki/Probability_density_function
    ys = ys if ys is not None else [1 for x in xs]
    assert all([0<=y for y in ys])
    assert np.allclose(ys[0],ys[-1],0), f"non-discrete PDF must be zero at ends: {ys[0]:g},{ys[-1]:g}"
    area = trapezoidintegrate(ys,xs,returnarray=True)[-1]
    invcdf = trapezoidintegrate(ys,xs,invfunc=True)
    if debug:
        from wavedata import Wave
        print(invcdf(0),invcdf(area))
        r = [ai/area for ai in trapezoidintegrate(ys,xs,returnarray=True)]
        w = Wave(r,xs).setplot(m='o',l='')
        rr = np.linspace(0,1,1001)
        ww = Wave(rr,[invcdf(ri*area) for ri in rr])
        Wave.plots(w,ww)
    return invcdf(np.random.random()*area)
def interpolate1d(x,xs,ys,kind='linear',extrapolate=None,checkvalidity=False):
    # kind: linear nearest nearest-up zero slinear quadratic cubic previous next # see scipy.interpolate.interp1d
    from scipy.interpolate import interp1d
    if len(xs)<2:
        return 0*x + ys[0]
    if not xs[0]<xs[-1]:
        xx = [x for x in xs if not np.isnan(x)]
        if not xx[0]<xx[-1]:
            xs,ys = xs[::-1],ys[::-1]
    def valid(xs): # 𝒪(1) monte carlo validity check
        from random import randrange
        i = randrange(0,len(xs)-1)
        # if not xs[i]<=xs[i+1]: print(i, xs[i], xs[i+1], xs[i]<=xs[i+1])
        return xs[i]<=xs[i+1] or np.isnan(xs[i]) or np.isnan(xs[i+1])
    def allvalid(xs): # 𝒪(N)
        return all([0<xi for xi in xs[1:]-xs[:-1]])
    assert valid(xs), 'failed monte carlo check, xs must be monotonically increasing or decreasing for interpolate1d (also use checkvalidity=True)'
    if checkvalidity: assert allvalid(xs),[xi for xi in xs[1:]-xs[:-1]]
    assert extrapolate in [None,'lin','linear','log','logarithmic','const','constant']
    if extrapolate in ['log','logarithmic']:
        ys = np.log(ys)
    fill_value = (ys[0],ys[-1]) if extrapolate in ['const','constant'] else (None if extrapolate is None else 'extrapolate')
    with np.errstate(invalid='ignore',divide='ignore'): # print(np.geterr())
        # suppress warnings when there are duplicate x values (eliminate dups instead?)
        f = interp1d(xs,ys,kind=kind,fill_value=fill_value,bounds_error=False)
        y = np.exp(f(x)) if extrapolate in ['log','logarithmic'] else f(x)
    # return y if hasattr(x,'__len__') else np.asscalar(y)
    return y if hasattr(x,'__len__') else y.item()
def fourierintegral(ys,xs,freq,norm=False,returnwaves=False): # exact fourier integral of piecewise linear function
    # @np.vectorize
    def integral(f):
        # wolfram alpha: integrate (u0 + (u1-u0)/(x1-x0)*(x-x0)) * cos(k x) from x0 to x1
        #   Integral(u(x)*cos(kx),x0,x1) = ((-u0 + u1) Cos[k x0] + (u0 - u1) Cos[k x1] - k (x0 - x1) (u0 Sin[k x0] - u1 Sin[k x1]))/(k^2 (x0 - x1)) =  (u1-u0)/(x1-x0)/k^2 * (cos(k*x1)-cos(k*x0)) + 1/k * (u1*sin(k*x1)-u0*sin(k*x0))
        # wolfram alpha: integrate (u0 + (u1-u0)/(x1-x0)*(x-x0)) * sin(k x) from x0 to x1
        #   Integral(u(x)*sin(kx),x0,x1) = (k u0 Cos[k x0] - k u1 Cos[k x1] - ((u0 - u1) (Sin[k x0] - Sin[k x1]))/(x0 - x1))/k^2 = -(u1-u0)/(x1-x0)/k^2 * (sin(k*x0)-sin(k*x1)) + 1/k * (u0*cos(k*x0)-u1*cos(k*x1))
        u0,u1,x0,x1 = np.array(ys[:-1]), np.array(ys[1:]), np.array(xs[:-1]), np.array(xs[1:])
        k = -2*pi*f # note negative sign to make consistent with np.fft # defines FT as ∫y(x)exp(-i2πxf)dx (instead of ∫y(x)exp(+i2πxf)dx)
        u0,u1,x0,x1 = u0[x1-x0!=0].astype(np.float64),u1[x1-x0!=0].astype(np.float64),x0[x1-x0!=0],x1[x1-x0!=0]
        # if 0==k: return ((u0+u1)/2*(x1-x0)).sum()
        wc =  (u1-u0)/(x1-x0)/k**2 * (cos(k*x1)-cos(k*x0)) + 1/k * (u1*sin(k*x1)-u0*sin(k*x0)) if not 0==k else ((u0+u1)/2*(x1-x0))
        ws = -(u1-u0)/(x1-x0)/k**2 * (sin(k*x0)-sin(k*x1)) + 1/k * (u0*cos(k*x0)-u1*cos(k*x1)) if not 0==k else 0*wc
        if norm:
            L = x1[-1]-x0[0]
            wc,ws = wc/L,ws/L
        if returnwaves:
            from wavedata import Wave
            wx = [x0[0]] + list(x1)
            wc,ws = Wave(lacc(wc,initial=0),wx), Wave(lacc(ws,initial=0),wx)
            return wc,ws
        return wc.sum()+1j*ws.sum()
    integral = integral if returnwaves else np.vectorize(integral)
    return integral(freq)
# def rms(x,power=2): # root-mean-square for power greater than two, without overflow in python, using np.logsumexp (from github copilot)
#     from scipy.special import logsumexp
#     return np.exp(logsumexp(np.log(np.abs(x)**power))/len(x))**(1./power)
def rms(numbers,power=2): # root-mean-square for power greater than two, without overflow in python, using np.logsumexp (from openai chat)
    from scipy.special import logsumexp
    log_sum_of_powers = logsumexp([power * np.log(np.abs(x)+1e-99) for x in numbers])
    return np.exp(log_sum_of_powers / power)
# def test_rms(): # (from openai chat)
#     assert np.isclose(rms([1,2,3,4],2),np.sqrt(sum([x**2 for x in [1,2,3,4]]))) # Test RMS for p = 2
#     assert np.isclose(rms([1,2,3,4],3),(sum([x**3 for x in [1,2,3,4]]))**(1/3.))  # Test RMS for p = 3
#     assert np.isclose(rms([1,2,3,4],4),(sum([x**4 for x in [1,2,3,4]]))**(1/4.))  # Test RMS for p = 4
# test_rms()
def multicrosscorrelation(ws,shifts=None,fraction=0.8,j0=0,di=None,normalize=True,plot=False,verbose=True):
    # an extension of the cross-correlation formula to more than two arrays (same as standard cross-correlation when len(ws)==2)
    # an extension of this formula in 1D to more than two arrays: https://youtu.be/1_hwFc8PXVE?t=350
    # - fraction is the minimum overlap betwen any two arrays
    #   it defines the maximum extent over which the correlation is computed (parts that don't overlap get clipped)
    # - j0 is the index of the ws array with respect to which all other arrays are shifted
    # - di is the shift resolution, di=1 is 1 pixel shift between first and second array (jj-1 shift between first and last)
    def rint(x): return int(round(x))
    from math import prod
    assert di is not None or shifts is not None
    jj,ii,ilen = len(ws),len(ws[0]),rint(fraction*len(ws[0]))
    i0,i1 = (ii-ilen)//2,(ii+ilen)//2 # range of the j0 array for all shifts
    assert i1-i0==ilen
    assert all(len(w)==ii for w in ws)
    assert 0<=j0<=jj-1
    def subrange(w,j,dx):
        n = rint(dx*(j-j0))
        # return w[i0+n:i1+n]
        return np.array(w[i0+n:i1+n],dtype=np.dtype('f8'))
    # n ranges from dx*(0-j0) to dx*(jj-1-j0) 
    # |n| < |dx|*(jj-1)
    dxmin,dxmax = -i0/(jj-1),+i0/(jj-1)
    ndx = int(dxmax/di)
    dxs = [n*di for n in range(-ndx,ndx+1)]
    assert len(dxs)==2*ndx+1
    dxs = shifts if shifts is not None else dxs
    def response(shift):
        from wavedata import Wave
        vs = [subrange(w,j,shift) for j,w in enumerate(ws)]
        if plot and shifts is not None: Wave.plots(*[Wave(v).rename(j) for j,v in enumerate(vs)],m=1,l='0123',ms=2)
        # norm = prod([sum(v**jj)**(1./jj) for v in vs]) if normalize else 1 # overflow error
        norm = prod([rms(v,power=jj) for v in vs]) if normalize else 1
        return sum(prod(vs))/norm
    a = np.array([response(shift) for shift in dxs])
    return a,dxs
def dotproduct(a,b):
    return sum([ai*bi for ai,bi in zip(a,b)])
def lerp(x,x0,x1,y0,y1):
    return y0 + (y1-y0)*(x-x0)/(x1-x0)
def polarlerp(φ,φ0,φ1,r0,r1,confine=True):
    # return r(φ) such that r is along straight line through (r0,φ0) and (r1,φ1)
    # if confine==True, return nan if not on segment connecting (r0,φ0) and (r1,φ1), e.g. when φ0<φ<φ1 and 0.5<φ1-φ0
    def h(z): # avoid divide by zero error
        return np.where(z,z,1e-19)
    x0,y0 = r0*cos(2*pi*φ0),r0*sin(2*pi*φ0)
    x1,y1 = r1*cos(2*pi*φ1),r1*sin(2*pi*φ1)
    m = (y1-y0)/h(x1-x0)
    c = np.where(φ0+0.5<φ,nan,0) + np.where(φ<φ1-0.5,nan,0) if confine else 0
    return c + (y0-m*x0)/h(sin(2*pi*φ)-m*cos(2*pi*φ))
def complex2rectangular(z):
    return np.real(z),np.imag(z)
def complex2polar(z):
    return np.abs(z),np.angle(z)
def endlessshuffle(n,choices,norepeats=True,seed=0): # returns nth result of endless shuffle, norepeats = never same one twice in a row
    import random
    k = len(choices)
    assert 2<k, 'endless shuffle requires deck size of 3 or more'
    def cycle(j): # indexes of jth shuffling cycle
        random.seed(j+seed)
        return random.sample(range(len(choices)), k=k)
    c0,c1 = [choices[i] for i in cycle(n//k)],[choices[i] for i in cycle(n//k+1)]
    if norepeats and c0[-1]==c1[0]:
        c0[-2],c0[-1] = c0[-1],c0[-2] # do a swap if last result of current cycle would equal the first of next (..abx,xcd.. becomes ..abx,cxd..)
    return c0[n%k], (c0+c1)[n%k+1] # return current choice, next choice
def defaultcolors():
    return [['#4C72B0','#55A868','#C44E52','#8172B2','#CCB974','#64B5CD'],['#3f3d99','#993d71','#998b3d','#3d9956','#3d5a99','#993d90','#996d3d','#43993d','#3d7999','#843d99','#994e3d','#62993d','#3d9799','#653d99','#993d4b'],['#575757','#a41b0b','#311b10','#cabb9c','#cf9807','#1b4129','#204f6d','#142551','#a89ab3','#856b74','#7c0818'],['#807f7f','#cb837f','#683e35','#cdaf9a','#493d33','#c1b493','#b2b990','#759381','#92a1ab','#575859','#435772'],['#9f2520','#a84e2c','#3e1e0f','#f1c0a2','#d67936','#d3b96f','#717f8c','#3f4f5f','#2b3d61','#4c2d32'],['#740d06','#ad5000','#e7a312','#2a6a1e','#daf5ff','#1f82bb','#a38fc0','#251d2a','#f00002'],['#681108','#dc6707','#efc600','#a3c1e5','#505ea5','#988dcf','#270d1a','#de88a3','#b8223f'],['#3b2d2a','#5d2f20','#88441d','#a48533','#dee3ad','#adc5ab','#69868c','#666769','#848283'],['#d2bbaf','#686d44','#9daf7c','#566f64','#869799','#6c757b','#647e91','#9b8f94'],['#be4d00','#e17e18','#d3cf74','#aeb0a5','#9cb057','#457357','#29475f','#26457b','#14295e','#a8a2ce'],['#bc7933','#81490c','#3f4c42','#658b6e','#8a80a2','#371e54','#5d4c70'],['#66c2a5','#fc8d62','#8da0cb','#e78ac3','#a6d854','#ffd92f','#e5c494','#b3b3b3'],['#e41a1c','#377eb8','#4daf4a','#984ea3','#ff7f00','#ffff33','#a65628','#f781bf','#999999'],['#405e73','#80bfe6','#2e8c2e','#8f8c45','#d9994d','#d9b366','#fadb96','#ffffff'],['#2e0053','#287c48','#4e1e58','#4a6c4a','#39ff00','#004a66','#002000','#002400','#007010','#48a230','#704c1c','#4e4e00','#00481f','#600028','#00264c','#002000','#eb4100','#084a00','#004a00','#4a0066','#300054','#536753','#536753','#536760','#f46076','#00674a','#66761c','#002f20','#2f2000','#004e3a','#004a67','#ff0030','#4a0067','#006720','#20002f','#2f002f','#4e2637','#005220','#2000be','#6dff70','#18ff4e','#010101','#004803','#002c00','#266c70','#006620','#4a0066','#006070','#2f2f4e','#3e4f00','#306028','#206653','#704c1c','#4e4e00','#fe481f','#002600','#002800','#002d00','#200024','#7c6000','#fe2f2f','#fe4e3e','#006600','#10fe0c','#660070','#fee020','#43fe20','#48fe48','#2f48fe','#004e0b','#204f00','#b7674a','#6d4a66','#fe200c','#b00066','#10fe0c','#6648ff','#fe2f4e','#3e4f00','#48ff48','#2f4e0b','#4f0066','#ff22ff','#20ff22','#230000','#52b86e','#704c1c','#4e4e00','#ff481f','#200026','#600120','#263c30','#546753','#536753','#536701','#670160','#f46001','#f46001','#4e0138','#660042','#002c00','#602000','#eb4300','#082dff','#00664a','#67287c','#005467','#675367','#675367','#67607a','#000060','#f46000','#52bc6e','#00004a','#702f2f','#002f00','#0f384f','#6738f4','#20002c','#7e6020','#20eb43','#41082d','#4a0066','#006728','#70b000','#390000','#52bc6e','#000053','#fe3060','#ff0041','#220070','#200020','#522020','#562224','#240070','#4aa220','#220022','#10d141','#220070','#d00052','#70a220','#54304c','#ff4e4e','#00023e','#004801','#007e60','#002000','#107832','#c26720','#267c20','#70b000','#30f560','#2000be','#6d704c','#ff4e4e','#010101','#ff480f','#2dff78','#002000','#4c70b0','#66004a','#6e70b0','#6d35ff','#422000','#007e60','#0020eb','#004708','#00664a','#67287c','#b00066','#b00066','#602052','#6e4a67','#002000','#702f2f','#002f00','#0d4f00','#205220','#b8006d','#20ffa0','#4c1cff','#4e0000','#ff4807','#007026','#7a2000','#200042','#007212','#4a6670','#662000','#002000','#002f00','#004e0f','#204f00','#2f002f','#4e183d','#506602','#6648ff','#48ff2f','#2f004e','#26204f','#670270','#662000','#002f4e','#285020','#200020','#c06770','#022000','#00672f','#2f4e02','#ff5066','#70b000','#701000','#00c267','#002f2f','#033dff','#006601','#600170','#663df4','#600170','#ff4aff','#48ff2f','#2f004e','#3dff4f','#66012c','#20e572','#2400e5','#2f4e25','#205866','#f5ff60','#290029','#007029','#2f004e','#290041','#430070','#70c058','#4a0067','#006019','#00416c','#004aff','#20ff21','#20004a','#205200','#004800','#002f2f','#0c2f00','#002f2f','#2f004e','#3dff4f','#66004a','#672600','#ff702f','#ff4e0e','#60fe20','#701000']]
def hms2sec(s):
    assert isinstance(s, str), f"{s} must be string, e.g. '1:34:03.89'"
    smh = s.split(':')[::-1]
    def fs(ss): return (float(ss[0]) if '.' in ss[0] else int(ss[0])) + 60*fs(ss[1:]) if ss else 0
    return fs(smh)
def sec2hms(tt,round=False):
    if tt<0:
        return '-'+sec2time(abs(tt)).strip()
    t = int(tt)
    h,m,s = t//3600,(t//60)%60,t%60
    hms = f'{h:02d}:{m:02d}:{s:02d}' if h else f'{m:2d}:{s:02d}'
    return hms if round else hms + f"{tt-t:.2f}"[1:]
def timeit(func): # decorator
    from time import time
    def f(*args, **kwargs):
        t0,result = time(),func(*args, **kwargs)
        print(f" {func.__name__ if hasattr(func,'__name__') else func.__class__.__name__}: {sec2hms(time()-t0)}")
        return result
    return f
def profile(func,ntop=20): # decorator
    def f(*args, **kwargs):
        import cProfile
        from pstats import Stats
        pr = cProfile.Profile()
        pr.enable()
        result = func(*args, **kwargs)
        pr.disable()
        Stats(pr).sort_stats('cumtime').print_stats(ntop)
        return result
    return f
class dotdict(dict): # dot.notation access to dictionary attributes
    __getattr__ = dict.get
    __setattr__ = dict.__setitem__
    __delattr__ = dict.__delitem__
